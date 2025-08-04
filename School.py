import tkinter as tk
from tkinter import ttk, messagebox
from datetime import date
import openpyxl
import os

DATA_FILE = 'student_data.xlsx'

# Ensure Excel file exists
def initialize_excel():
    if not os.path.exists(DATA_FILE):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        headers = ["Registration No", "Name", "Class", "Gender", "DOB", "Date of Registration", "Fees Paid", "Exam Timetable", "Assessments"]
        sheet.append(headers)
        workbook.save(DATA_FILE)

# Add student to Excel
def save_student():
    reg_no = reg_var.get()
    name = name_var.get()
    student_class = class_var.get()
    gender = gender_var.get()
    dob = dob_var.get()
    fees = fees_var.get()
    exam_tt = exam_tt_var.get()
    assessment = assessment_var.get()
    date_of_reg = date.today().strftime("%d/%m/%Y")

    if not (reg_no and name and student_class and gender and dob):
        messagebox.showerror("Input Error", "All fields are required!")
        return

    workbook = openpyxl.load_workbook(DATA_FILE)
    sheet = workbook.active
    sheet.append([reg_no, name, student_class, gender, dob, date_of_reg, fees, exam_tt, assessment])
    workbook.save(DATA_FILE)
    messagebox.showinfo("Success", "Student data saved successfully!")
    clear_form()

# Clear form fields
def clear_form():
    reg_var.set("")
    name_var.set("")
    class_var.set("")
    gender_var.set("")
    dob_var.set("")
    fees_var.set("")
    exam_tt_var.set("")
    assessment_var.set("")

# View student data
def view_students():
    top = tk.Toplevel(root)
    top.title("Student Records")
    top.geometry("900x400")

    tree = ttk.Treeview(top, columns=("RegNo", "Name", "Class", "Gender", "DOB", "DateReg", "Fees", "ExamTT", "Assessment"), show='headings')
    tree.heading("RegNo", text="Registration No")
    tree.heading("Name", text="Name")
    tree.heading("Class", text="Class")
    tree.heading("Gender", text="Gender")
    tree.heading("DOB", text="DOB")
    tree.heading("DateReg", text="Date of Registration")
    tree.heading("Fees", text="Fees Paid")
    tree.heading("ExamTT", text="Exam Timetable")
    tree.heading("Assessment", text="Assessments")

    tree.pack(fill=tk.BOTH, expand=True)

    workbook = openpyxl.load_workbook(DATA_FILE)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        tree.insert('', tk.END, values=row)

# Dashboard
def dashboard():
    dash = tk.Toplevel(root)
    dash.title("Dashboard")
    dash.geometry("300x300")
    tk.Label(dash, text="Student Management Dashboard", font=('Arial', 14)).pack(pady=10)
    tk.Button(dash, text="Add Student", command=root.deiconify).pack(pady=5)
    tk.Button(dash, text="View Students", command=view_students).pack(pady=5)

# Main Application Window
root = tk.Tk()
root.title("Student Management System")
root.geometry("500x600")

initialize_excel()

# Variables
reg_var = tk.StringVar()
name_var = tk.StringVar()
class_var = tk.StringVar()
gender_var = tk.StringVar()
dob_var = tk.StringVar()
fees_var = tk.StringVar()
exam_tt_var = tk.StringVar()
assessment_var = tk.StringVar()

# Form Labels and Entries
tk.Label(root, text="Registration No:").pack(pady=5)
tk.Entry(root, textvariable=reg_var).pack(pady=5)

tk.Label(root, text="Name:").pack(pady=5)
tk.Entry(root, textvariable=name_var).pack(pady=5)

tk.Label(root, text="Class:").pack(pady=5)
tk.Entry(root, textvariable=class_var).pack(pady=5)

tk.Label(root, text="Gender:").pack(pady=5)
tk.Entry(root, textvariable=gender_var).pack(pady=5)

tk.Label(root, text="Date of Birth (dd/mm/yyyy):").pack(pady=5)
tk.Entry(root, textvariable=dob_var).pack(pady=5)

tk.Label(root, text="Fees Paid:").pack(pady=5)
tk.Entry(root, textvariable=fees_var).pack(pady=5)

tk.Label(root, text="Exam Timetable:").pack(pady=5)
tk.Entry(root, textvariable=exam_tt_var).pack(pady=5)

tk.Label(root, text="Assessments:").pack(pady=5)
tk.Entry(root, textvariable=assessment_var).pack(pady=5)

# Buttons
tk.Button(root, text="Save Student", command=save_student).pack(pady=10)
tk.Button(root, text="View Students", command=view_students).pack(pady=5)
tk.Button(root, text="Clear Form", command=clear_form).pack(pady=5)
tk.Button(root, text="Dashboard", command=dashboard).pack(pady=10)

root.mainloop()
